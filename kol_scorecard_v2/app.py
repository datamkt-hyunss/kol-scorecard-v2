"""
KOL Scorecard v2 — Streamlit App
가이드 v4 기준 | 멀티플랫폼 | 국가별 환율·기준치 | 세션 쿠키 로그인
"""
import subprocess, sys, io, re
import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from scoring import evaluate_batch, get_grade, stars_from_score, CUTOFFS, WEIGHTS

# ── Playwright 자동 설치 ──────────────────────────────────
@st.cache_resource(show_spinner=False)
def _install_playwright():
    try:
        subprocess.run([sys.executable,"-m","playwright","install","chromium"],
                       capture_output=True, timeout=180)
    except Exception:
        pass

_install_playwright()

# ── 페이지 설정 ───────────────────────────────────────────
st.set_page_config(page_title="KOL Scorecard v2", page_icon="📊",
                   layout="wide", initial_sidebar_state="expanded")

st.markdown("""
<style>
* { font-family: 'Noto Sans JP', 'Malgun Gothic', Arial, sans-serif !important; }
.sh { background:linear-gradient(135deg,#1F4E79,#2E75B6);color:#fff;
      border-radius:8px;padding:10px 16px;font-weight:700;font-size:1.05em;margin:8px 0 6px; }
.card { border:1px solid #e0e0e0;border-radius:10px;padding:14px 18px;
        margin-bottom:10px;background:#fff; }
.badge-adopt  { background:#1D6A2D;color:#fff;border-radius:5px;padding:2px 9px;font-weight:700;font-size:0.9em; }
.badge-review { background:#DCB306;color:#fff;border-radius:5px;padding:2px 9px;font-weight:700;font-size:0.9em; }
.badge-reject { background:#7B0000;color:#fff;border-radius:5px;padding:2px 9px;font-weight:700;font-size:0.9em; }
.login-box { background:#F0F7FF;border:1px solid #BDD7EE;border-radius:8px;padding:14px;margin-bottom:8px; }
.step-label { background:#2E75B6;color:#fff;border-radius:20px;padding:2px 12px;
              font-size:0.85em;font-weight:700;display:inline-block;margin-bottom:8px; }
div[data-testid="stDataFrame"] table td, div[data-testid="stDataFrame"] table th {
    font-size:0.82em !important; }
</style>
""", unsafe_allow_html=True)

RATE_DEFAULT = 9.5

# ── 플랫폼 정의 ───────────────────────────────────────────
PLATFORMS = {
    "tiktok":    {"label": "TikTok",           "icon": "🎵", "needs_login": True},
    "ig_feed":   {"label": "Instagram 피드",    "icon": "📸", "needs_login": True},
    "ig_reels":  {"label": "Instagram 릴스",    "icon": "🎬", "needs_login": True},
    "youtube":   {"label": "YouTube",           "icon": "▶️", "needs_login": False},
}

PLATFORM_COLS = {
    "tiktok":   ["KOL명","URL","비용","플랫폼","핀게시물ID"],
    "ig_feed":  ["KOL명","URL","비용","플랫폼","핀게시물ID"],
    "ig_reels": ["KOL명","URL","비용","플랫폼","핀게시물ID"],
    "youtube":  ["KOL명","URL","비용","플랫폼"],
}

# ═══════════════════════════════════════════════════════════
# 사이드바
# ═══════════════════════════════════════════════════════════
with st.sidebar:
    st.title("📊 KOL Scorecard v2")
    st.caption("가이드 v4 | 멀티플랫폼 | 국가별 기준")
    st.divider()

    # ── 국가 / 환율 설정 ────────────────────────────────
    st.subheader("🌏 국가 & 환율 설정")
    country = st.selectbox("평가 기준 국가", ["일본 (JPY)", "한국 (KRW)", "미국 (USD)"])
    if "일본" in country:
        currency = "JPY"
        rate_label = "¥100 = ₩?"
        rate_default = 950
    elif "한국" in country:
        currency = "KRW"
        rate_label = "기준 통화 (₩)"
        rate_default = 1
    else:
        currency = "USD"
        rate_label = "$1 = ₩?"
        rate_default = 1350

    rate_input = st.number_input(rate_label, value=rate_default, step=10)
    if currency == "JPY":
        rate = rate_input / 100
    elif currency == "USD":
        rate = rate_input
    else:
        rate = 1.0

    st.divider()

    # ── 커스텀 컷오프 설정 ──────────────────────────────
    st.subheader("📐 컷오프 기준")
    use_custom = st.checkbox("커스텀 기준치 사용", value=False)

    if use_custom:
        st.caption("TikTok 기준치")
        custom_cpv_e = st.number_input("CPV 탁월 기준 (미만 ₩)", value=12, step=1)
        custom_cpv_g = st.number_input("CPV 양호 기준 (미만 ₩)", value=23, step=1)
        custom_er_e  = st.number_input("ER% 탁월 기준 (이상 %)", value=3.48, step=0.1, format="%.2f")
        CUTOFFS["tiktok"]["cpv"]["E"] = custom_cpv_e
        CUTOFFS["tiktok"]["cpv"]["G"] = custom_cpv_g
        CUTOFFS["tiktok"]["er"]["E"]  = custom_er_e

    with st.expander("현재 TikTok 컷오프"):
        tk_c = CUTOFFS["tiktok"]
        st.table(pd.DataFrame({
            "지표":["CPV(₩)","ER%","저장률%","공유율%"],
            "◎탁월":["<"+str(tk_c['cpv']['E']),"≥"+str(tk_c['er']['E'])+"%",
                     "≥"+str(tk_c['save']['E'])+"%","≥"+str(tk_c['share']['E'])+"%"],
            "✕저조":["≥"+str(tk_c['cpv']['C']),"<"+str(tk_c['er']['C'])+"%",
                     "<"+str(tk_c['save']['C'])+"%","<"+str(tk_c['share']['C'])+"%"],
        }))

    st.divider()
    n_posts = st.slider("스크래핑 게시물 수", 5, 15, 10, 1)
    st.divider()

    # ── 로그인 / 세션쿠키 ────────────────────────────────
    st.subheader("🔑 플랫폼별 로그인")
    st.markdown('<div class="login-box">', unsafe_allow_html=True)
    st.caption("☁️ 클라우드 로그인 방법\n① 아래 버튼으로 플랫폼에 로그인\n② 브라우저 개발자도구(F12) → Application → Cookies\n③ sessionid 값을 복사해서 아래에 붙여넣기")
    st.markdown('</div>', unsafe_allow_html=True)

    if "sessions" not in st.session_state:
        st.session_state.sessions = {"tiktok":"","instagram":""}

    st.markdown("**TikTok 세션쿠키**")
    tk_session = st.text_input("TikTok sessionid", value=st.session_state.sessions.get("tiktok",""),
                                type="password", placeholder="복사한 sessionid 값 붙여넣기", key="tk_sid")
    st.session_state.sessions["tiktok"] = tk_session

    st.markdown("**Instagram 세션쿠키**")
    ig_session = st.text_input("Instagram sessionid", value=st.session_state.sessions.get("instagram",""),
                                type="password", placeholder="복사한 sessionid 값 붙여넣기", key="ig_sid")
    st.session_state.sessions["instagram"] = ig_session

    st.divider()
    st.subheader("결론 로직 v4")
    st.info("**1차:** 절대 < 5.0 → 제외 ✗\n\n**예외:** 절대 ≥ 8.0 & 상대 < 4.0 → 검토 △\n\n**2차:** 상대 ≥ 6.0 → 채택 ✓\n4.0~5.9 → 검토 △ | < 4.0 → 제외 ✗")


# ═══════════════════════════════════════════════════════════
# Excel 출력 (FIN 파일 동일 양식)
# ═══════════════════════════════════════════════════════════
def make_excel(tk_results, ig_results):
    wb = openpyxl.Workbook()
    def P(h): return PatternFill('solid', fgColor='FF'+h if len(h)==6 else h)
    def F(bold=False,sz=9,color='000000',italic=False):
        return Font(name='Arial',bold=bold,size=sz,color=color,italic=italic)
    thin = Side(border_style='thin',color='BFBFBF')
    BALL = Border(left=thin,right=thin,top=thin,bottom=thin)
    AC=Alignment(horizontal='center',vertical='center',wrap_text=True)
    AL=Alignment(horizontal='left',vertical='center',wrap_text=False)
    ALW=Alignment(horizontal='left',vertical='center',wrap_text=True)
    FILLS={
        'hdr_title':P('1F4E79'),'hdr_blue':P('2E75B6'),'hdr_green':P('375623'),
        'hdr_navy':P('1F4E79'),'hdr_purple':P('4A235A'),'hdr_light':P('BDD7EE'),
        'hdr_grade':P('C6E0B4'),'hdr_result':P('D8B4FE'),'row_odd':P('F7FBFF'),
        'row_even':P('FFFFFF'),'rel_score':P('EEF4FB'),'note_bg':P('F2F2F2'),
        'cmt_bg':P('FFF2CC'),'E':P('E2EFDA'),'G':P('DDEBF7'),
        'C':P('FFF2CC'),'H':P('FCE4D6'),'adopt':P('1D6A2D'),'review':P('DCB306'),'reject':P('7B0000'),
    }
    GF={'◎ 탁월':F(True,9,'375623'),'○ 양호':F(True,9,'1F4E79'),'△ 보통':F(True,9,'7D6608'),'✕ 저조':F(True,9,'843C0C')}
    GFL={'◎ 탁월':FILLS['E'],'○ 양호':FILLS['G'],'△ 보통':FILLS['C'],'✕ 저조':FILLS['H']}
    CF={'채택 ✓':FILLS['adopt'],'검토 △':FILLS['review'],'제외 ✗':FILLS['reject']}
    def cw(ws,c,w): ws.column_dimensions[get_column_letter(c)].width=w
    def rh(ws,r,h): ws.row_dimensions[r].height=h
    def mg(ws,rng): ws.merge_cells(rng)
    def hc(ws,r,c,v,fill,font=None,al=AC):
        cc=ws.cell(r,c,v); cc.fill=fill; cc.font=font or F(True,9,'FFFFFF'); cc.alignment=al; cc.border=BALL
    def dc(ws,r,c,v,fill,font=None,al=AC,fmt=None):
        cc=ws.cell(r,c,v); cc.fill=fill; cc.font=font or F(False,9); cc.alignment=al; cc.border=BALL
        if fmt: cc.number_format=fmt
    def gc(ws,r,c,g):
        cc=ws.cell(r,c,g); cc.fill=GFL.get(g,FILLS['row_even']); cc.font=GF.get(g,F(True,9)); cc.alignment=AC; cc.border=BALL
    def rc(ws,r,c,v):
        cc=ws.cell(r,c,round(v,2)); cc.fill=FILLS['rel_score']; cc.font=F(True,10); cc.alignment=AC; cc.border=BALL; cc.number_format='0.00'
    def cc_(ws,r,c,con):
        cc=ws.cell(r,c,con); cc.fill=CF.get(con,FILLS['reject']); cc.font=F(True,10,'FFFFFF'); cc.alignment=AC; cc.border=BALL
    def cmt(ws,r,c,t,rf):
        cc=ws.cell(r,c,t or ''); cc.fill=FILLS['cmt_bg'] if t else rf
        cc.font=F(False,9,'7D6608') if t else F(False,9); cc.alignment=ALW; cc.border=BALL
    def note(ws,r,c1,c2,t):
        mg(ws,f'{get_column_letter(c1)}{r}:{get_column_letter(c2)}{r}')
        cc=ws.cell(r,c1,t); cc.fill=FILLS['note_bg']; cc.font=F(False,8,'595959',True); cc.alignment=AL

    # TikTok 시트
    if tk_results:
        ws=wb.active; ws.title='TikTok'; ws.sheet_view.showGridLines=False; ws.freeze_panes='A4'
        for i,w in enumerate([5,16,10,10,8,6,8,7,9,8,7,7,7,9,9,9,9,9,8,8,28],1): cw(ws,i,w)
        rh(ws,1,18); rh(ws,2,30); rh(ws,3,31.9)
        mg(ws,'A1:U1'); hc(ws,1,1,'TikTok KOL 스코어 평가표  |  기준: KOL 선별 프레임워크 가이드 v4  |  환율 적용',FILLS['hdr_title'],F(True,11,'FFFFFF'),AL)
        for rng,lbl,fill in [('A2:B2','기본 정보',FILLS['hdr_blue']),('C2:H2','원본 데이터',FILLS['hdr_blue']),('I2:M2','산출 지표 (₩)',FILLS['hdr_blue']),('N2:Q2','지표별 등급',FILLS['hdr_blue']),('R2','절대 종합점수',FILLS['hdr_green']),('S2','상대 종합점수',FILLS['hdr_navy']),('T2:U2','최종 평가',FILLS['hdr_purple'])]:
            mg(ws,rng); cc2=ws[rng.split(':')[0]]; cc2.value=lbl; cc2.fill=fill; cc2.font=F(True,9,'FFFFFF'); cc2.alignment=AC
        for col,val,fk in [(1,'No.','hdr_light'),(2,'KOL명','hdr_light'),(3,'비용','hdr_light'),(4,'조회수','hdr_light'),(5,'좋아요','hdr_light'),(6,'댓글','hdr_light'),(7,'저장','hdr_light'),(8,'공유','hdr_light'),(9,'비용(₩)','hdr_light'),(10,'CPV(₩)','hdr_light'),(11,'ER%','hdr_light'),(12,'저장률%','hdr_light'),(13,'공유율%','hdr_light'),(14,'CPV\n등급','hdr_light'),(15,'ER\n등급','hdr_light'),(16,'저장률\n등급','hdr_light'),(17,'공유율\n등급','hdr_light'),(18,'절대\n등급','hdr_grade'),(19,'상대\n점수','hdr_light'),(20,'결론','hdr_result'),(21,'코멘트','hdr_result')]:
            hc(ws,3,col,val,FILLS[fk],F(True,8,'1F4E79'))
        for idx,rd in enumerate(tk_results):
            r=idx+4; rf=FILLS['row_odd'] if idx%2==0 else FILLS['row_even']; rh(ws,r,19.9); m=rd.metrics
            dc(ws,r,1,idx+1,rf,F(True,9)); dc(ws,r,2,rd.name,rf,F(True,9),AL)
            dc(ws,r,3,int(rd.cost_jpy),rf,fmt='#,##0'); dc(ws,r,4,int(m.get('_views',0)),rf,fmt='#,##0')
            dc(ws,r,5,int(m.get('_likes',0)),rf,fmt='#,##0'); dc(ws,r,6,int(m.get('_comments',0)),rf,fmt='#,##0')
            dc(ws,r,7,int(m.get('_saves',0)),rf,fmt='#,##0'); dc(ws,r,8,int(m.get('_shares',0)),rf,fmt='#,##0')
            dc(ws,r,9,int(m.get('_cost_krw',rd.cost_jpy*rate)),rf,fmt='#,##0'); dc(ws,r,10,round(m.get('cpv',0),1),rf,fmt='#,##0.0')
            dc(ws,r,11,round(m.get('er',0),3),rf,fmt='0.000'); dc(ws,r,12,round(m.get('save',0),3),rf,fmt='0.000'); dc(ws,r,13,round(m.get('share',0),3),rf,fmt='0.000')
            gc(ws,r,14,rd.grades.get('cpv','✕ 저조')); gc(ws,r,15,rd.grades.get('er','✕ 저조'))
            gc(ws,r,16,rd.grades.get('save','✕ 저조')); gc(ws,r,17,rd.grades.get('share','✕ 저조'))
            gc(ws,r,18,rd.abs_grade); rc(ws,r,19,rd.rel_score); cc_(ws,r,20,rd.conclusion); cmt(ws,r,21,rd.comment,rf)
        nr=len(tk_results)+5
        note(ws,nr,1,21,'【컷오프】 CPV: E<12 / G 12~23 / C 23~29 / H≥29  |  ER%: E≥3.48% / G 1.85% / C 1.05%  |  저장률%: E≥0.490% / G 0.230% / C 0.122%  |  공유율%: E≥0.058% / G 0.024% / C 0.013%')
        note(ws,nr+1,1,21,'【가중치】 CPV 30% / ER 30% / 저장률 25% / 공유율 15%  |  【결론 v4】 절대<5→제외 / 절대≥8&상대<4→검토 / 상대≥6→채택 / 4~5.9→검토 / <4→제외')
    else:
        ws=wb.active; ws.title='TikTok'; ws.cell(1,1,"데이터 없음")

    # IG 피드 시트
    ws2=wb.create_sheet('Instagram 피드')
    ws2.sheet_view.showGridLines=False; ws2.freeze_panes='A4'
    if ig_results:
        for i,w in enumerate([5,16,10,8,6,9,8,10,10,11,12,9,8,8,28],1): cw(ws2,i,w)
        rh(ws2,1,18); rh(ws2,2,30); rh(ws2,3,31.9)
        mg(ws2,'A1:O1'); hc(ws2,1,1,'Instagram 피드 KOL 스코어 평가표  |  기준: KOL 선별 프레임워크 가이드 v4',FILLS['hdr_title'],F(True,11,'FFFFFF'),AL)
        for rng,lbl,fill in [('A2:B2','기본 정보',FILLS['hdr_blue']),('C2:E2','원본 데이터',FILLS['hdr_blue']),('F2:I2','산출 지표 (₩)',FILLS['hdr_blue']),('J2:K2','지표별 등급',FILLS['hdr_blue']),('L2','절대 종합점수',FILLS['hdr_green']),('M2','상대 종합점수',FILLS['hdr_navy']),('N2:O2','최종 평가',FILLS['hdr_purple'])]:
            mg(ws2,rng); cc2=ws2[rng.split(':')[0]]; cc2.value=lbl; cc2.fill=fill; cc2.font=F(True,9,'FFFFFF'); cc2.alignment=AC
        for col,val,fk in [(1,'No.','hdr_light'),(2,'KOL명','hdr_light'),(3,'비용','hdr_light'),(4,'좋아요','hdr_light'),(5,'댓글','hdr_light'),(6,'비용(₩)','hdr_light'),(7,'총 참여수','hdr_light'),(8,'CPE(₩)','hdr_light'),(9,'댓글·\n좋아요 비율','hdr_light'),(10,'CPE\n등급','hdr_light'),(11,'댓글·좋아요\n등급','hdr_light'),(12,'절대\n등급','hdr_grade'),(13,'상대\n점수','hdr_light'),(14,'결론','hdr_result'),(15,'코멘트','hdr_result')]:
            hc(ws2,3,col,val,FILLS[fk],F(True,8,'1F4E79'))
        for idx,rd in enumerate(ig_results):
            r=idx+4; rf=FILLS['row_odd'] if idx%2==0 else FILLS['row_even']; rh(ws2,r,19.9); m=rd.metrics
            dc(ws2,r,1,idx+1,rf,F(True,9)); dc(ws2,r,2,rd.name,rf,F(True,9),AL)
            dc(ws2,r,3,int(rd.cost_jpy),rf,fmt='#,##0'); dc(ws2,r,4,int(m.get('_likes',0)),rf,fmt='#,##0'); dc(ws2,r,5,int(m.get('_comments',0)),rf,fmt='#,##0')
            dc(ws2,r,6,int(m.get('_cost_krw',rd.cost_jpy*rate)),rf,fmt='#,##0'); dc(ws2,r,7,int(m.get('_total',0)),rf,fmt='#,##0')
            dc(ws2,r,8,int(m.get('cpe',0)),rf,fmt='#,##0'); dc(ws2,r,9,round(m.get('clr',0),4),rf,fmt='0.0000')
            gc(ws2,r,10,rd.grades.get('cpe','✕ 저조')); gc(ws2,r,11,rd.grades.get('clr','✕ 저조'))
            gc(ws2,r,12,rd.abs_grade); rc(ws2,r,13,rd.rel_score); cc_(ws2,r,14,rd.conclusion); cmt(ws2,r,15,rd.comment,rf)
        ni=len(ig_results)+5
        note(ws2,ni,1,15,'【컷오프】 CPE: E<1,631 / G 1,631~3,465 / C 3,465~6,261 / H≥6,261  |  댓글·좋아요비율: E≥0.0042 / G 0.0002~0.0042 / C~0.0000  |  가중치: CPE 60% / CLR 40%')
        note(ws2,ni+1,1,15,'【결론 v4】 절대<5→제외 / 절대≥8&상대<4→검토+코멘트 / 상대≥6→채택 / 4~5.9→검토 / <4→제외')
    else:
        ws2.cell(1,1,"데이터 없음")

    buf=io.BytesIO(); wb.save(buf); return buf.getvalue()


# ═══════════════════════════════════════════════════════════
# 공통: 결과 표시
# ═══════════════════════════════════════════════════════════
CON_BG = {'채택 ✓':'#E8F5E9','검토 △':'#FFFDE7','제외 ✗':'#FFEBEE'}

def show_results(results_by_platform: dict):
    all_r = [r for rs in results_by_platform.values() for r in rs]
    if not all_r: return
    st.divider()
    st.markdown('<div class="sh">📊 평가 결과</div>', unsafe_allow_html=True)
    total=len(all_r); adopt=sum(1 for r in all_r if r.conclusion=='채택 ✓')
    review=sum(1 for r in all_r if r.conclusion=='검토 △'); reject=total-adopt-review
    c1,c2,c3,c4=st.columns(4)
    c1.metric("전체",f"{total}명"); c2.metric("채택 ✓",f"{adopt}명",f"{adopt/total*100:.0f}%" if total else "")
    c3.metric("검토 △",f"{review}명",f"{review/total*100:.0f}%" if total else "")
    c4.metric("제외 ✗",f"{reject}명",f"{reject/total*100:.0f}%" if total else "")

    tabs_labels=[f"{PLATFORMS[p]['icon']} {PLATFORMS[p]['label']}" for p in results_by_platform if results_by_platform[p]] + ["📥 Excel"]
    tabs_keys=[p for p in results_by_platform if results_by_platform[p]] + ["excel"]

    if len(tabs_labels) == 1:
        tabs = st.tabs(tabs_labels)
        tab_list = tabs
    else:
        tab_list = st.tabs(tabs_labels)

    for i,(tab,key) in enumerate(zip(tab_list, tabs_keys)):
        with tab:
            if key == "excel":
                tk_r=results_by_platform.get("tiktok",[])
                ig_r=results_by_platform.get("ig_feed",[])+results_by_platform.get("ig_reels",[])
                if tk_r or ig_r:
                    xl=make_excel(tk_r,ig_r)
                    st.download_button("📥 Excel 다운로드 (FIN 양식)",data=xl,
                        file_name="KOL_스코어_재평가_결과_v4.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary",use_container_width=True)
                    st.caption("TikTok 시트 + Instagram 피드 시트 포함. FIN 파일 동일 양식.")
            else:
                rs=results_by_platform[key]
                rows=[]
                for r in sorted(rs,key=lambda x:x.rel_score,reverse=True):
                    m=r.metrics
                    row={"KOL명":r.name,"절대등급":r.abs_grade,"상대점수":f"{r.rel_score:.2f}","결론":r.conclusion}
                    if key=="tiktok":
                        row.update({"CPV(₩)":f"{m.get('cpv',0):.1f}","ER%":f"{m.get('er',0):.3f}",
                                    "저장률%":f"{m.get('save',0):.3f}","공유율%":f"{m.get('share',0):.3f}"})
                    elif key in ("ig_feed","ig_reels"):
                        row.update({"CPE(₩)":f"{int(m.get('cpe',0)):,}","CLR":f"{m.get('clr',0):.4f}"})
                    if r.comment: row["코멘트"]=r.comment
                    rows.append(row)
                df=pd.DataFrame(rows)
                def style_fn(row):
                    bg=CON_BG.get(row["결론"],"#fff")
                    return [f"background-color:{bg}"]*len(row)
                st.dataframe(df.style.apply(style_fn,axis=1),use_container_width=True,height=min(60+len(df)*36,480))


# ═══════════════════════════════════════════════════════════
# 메인 탭
# ═══════════════════════════════════════════════════════════
tab_scrape, tab_manual, tab_guide = st.tabs(["🤖 자동 스크래핑", "📝 수동 입력", "📖 가이드 v4"])


# ─── 자동 스크래핑 ────────────────────────────────────────
with tab_scrape:
    st.markdown('<div class="sh">🤖 자동 스크래핑 & 평가</div>', unsafe_allow_html=True)

    col_mode,col_posts=st.columns([2,1])
    test_mode=col_mode.checkbox("🧪 테스트 모드 (더미 데이터로 로직 확인)", value=False,
                                 help="ON: 실제 스크래핑 없이 더미 데이터로 평가 로직 테스트")
    if test_mode:
        st.info("테스트 모드 ON — 더미 데이터로 평가 결과를 확인합니다. 실제 스크래핑 시에는 체크 해제하세요.")

    st.divider()

    # ── STEP 1: KOL 정보 입력 ─────────────────────────────
    st.markdown('<span class="step-label">STEP 1</span> **KOL 정보 입력**', unsafe_allow_html=True)

    input_method=st.radio("입력 방식",["✏️ 직접 입력","📂 CSV/Excel 업로드"],horizontal=True,label_visibility="collapsed")

    # 초기 데이터
    if "kol_df" not in st.session_state:
        st.session_state.kol_df = pd.DataFrame(columns=["KOL명","URL (필수)","플랫폼","캐스팅 비용(JPY)","핀 게시물 ID"])

    if "✏️" in input_method:
        st.caption("URL을 입력하면 플랫폼이 자동 감지됩니다.")

        col_add,col_init,col_sample,_=st.columns([1,1,1,5])
        if col_add.button("➕ 행 추가"):
            new_row=pd.DataFrame([{"KOL명":"","URL (필수)":"","플랫폼":"자동감지","캐스팅 비용(JPY)":0,"핀 게시물 ID":""}])
            st.session_state.kol_df=pd.concat([st.session_state.kol_df,new_row],ignore_index=True)
        if col_init.button("🗑️ 초기화"):
            st.session_state.kol_df=pd.DataFrame(columns=["KOL명","URL (필수)","플랫폼","캐스팅 비용(JPY)","핀 게시물 ID"])
        if col_sample.button("📋 샘플 데이터"):
            st.session_state.kol_df=pd.DataFrame([
                {"KOL명":"가연がよん","URL (필수)":"https://www.tiktok.com/@kayeon_japan","플랫폼":"자동감지","캐스팅 비용(JPY)":750000,"핀 게시물 ID":""},
                {"KOL명":"𝗸𝗮𝗻𝗮𝗻","URL (필수)":"https://www.tiktok.com/@kanan_official","플랫폼":"자동감지","캐스팅 비용(JPY)":150000,"핀 게시물 ID":""},
                {"KOL명":"NoChi","URL (필수)":"https://www.instagram.com/nochi_official/","플랫폼":"자동감지","캐스팅 비용(JPY)":460000,"핀 게시물 ID":""},
            ])

        # 플랫폼 자동 감지
        def detect_platform(url):
            if not url: return "자동감지"
            url=url.lower()
            if "tiktok.com" in url: return "TikTok"
            if "instagram.com" in url:
                return "Instagram 릴스" if "/reel" in url else "Instagram 피드"
            if "youtube.com" in url or "youtu.be" in url:
                return "YouTube 숏츠" if "/shorts" in url else "YouTube"
            return "자동감지"

        edited=st.data_editor(
            st.session_state.kol_df,
            use_container_width=True,
            num_rows="dynamic",
            column_config={
                "KOL명": st.column_config.TextColumn("KOL명",width="medium"),
                "URL (필수)": st.column_config.TextColumn("URL (필수)",width="large"),
                "플랫폼": st.column_config.SelectboxColumn("플랫폼",
                    options=["자동감지","TikTok","Instagram 피드","Instagram 릴스","YouTube 숏츠","YouTube"],
                    width="small"),
                "캐스팅 비용(JPY)": st.column_config.NumberColumn("캐스팅 비용(JPY)",format="¥%d",width="small"),
                "핀 게시물 ID": st.column_config.TextColumn("핀 게시물 ID",width="small"),
            },
            hide_index=True,
            key="kol_editor"
        )
        # 플랫폼 자동감지 적용
        if edited is not None:
            for idx,row in edited.iterrows():
                if str(row.get("플랫폼","")) in ["자동감지",""]:
                    detected=detect_platform(str(row.get("URL (필수)","") or ""))
                    if detected != "자동감지":
                        edited.at[idx,"플랫폼"]=detected
            st.session_state.kol_df=edited

    else:  # CSV/Excel 업로드
        st.caption("KOL명, URL, 비용, 플랫폼 컬럼이 포함된 파일을 업로드하세요.")
        uploaded=st.file_uploader("CSV 또는 Excel 파일",type=["csv","xlsx","xls"])
        if uploaded:
            try:
                if uploaded.name.endswith(".csv"):
                    df_up=pd.read_csv(uploaded)
                else:
                    df_up=pd.read_excel(uploaded)
                # 컬럼명 유연하게 매핑
                col_map={}
                for c in df_up.columns:
                    cl=c.lower().replace(" ","")
                    if "kol" in cl or "이름" in cl or "name" in cl: col_map[c]="KOL명"
                    elif "url" in cl or "링크" in cl: col_map[c]="URL (필수)"
                    elif "비용" in cl or "cost" in cl or "fee" in cl or "price" in cl: col_map[c]="캐스팅 비용(JPY)"
                    elif "플랫폼" in cl or "platform" in cl: col_map[c]="플랫폼"
                df_up=df_up.rename(columns=col_map)
                for req in ["KOL명","URL (필수)"]:
                    if req not in df_up.columns: df_up[req]=""
                if "캐스팅 비용(JPY)" not in df_up.columns: df_up["캐스팅 비용(JPY)"]=0
                if "플랫폼" not in df_up.columns: df_up["플랫폼"]="자동감지"
                if "핀 게시물 ID" not in df_up.columns: df_up["핀 게시물 ID"]=""
                st.session_state.kol_df=df_up[["KOL명","URL (필수)","플랫폼","캐스팅 비용(JPY)","핀 게시물 ID"]]
                st.success(f"{len(df_up)}개 KOL 로드 완료")
                st.dataframe(st.session_state.kol_df,use_container_width=True,height=200)
            except Exception as e:
                st.error(f"파일 읽기 오류: {e}")

    st.divider()

    # ── STEP 2: 스크래핑 실행 ─────────────────────────────
    st.markdown('<span class="step-label">STEP 2</span> **스크래핑 & 평가 실행**', unsafe_allow_html=True)

    if st.button("🚀 스크래핑 시작", type="primary", use_container_width=True):
        df_kol = st.session_state.kol_df.copy()
        df_kol = df_kol[df_kol["URL (필수)"].notna() & (df_kol["URL (필수)"]!="")].reset_index(drop=True)

        if df_kol.empty:
            st.warning("URL을 입력하세요.")
        else:
            import random as _r

            def detect_plat(url, stated):
                if stated and stated not in ["자동감지",""]:
                    m={"TikTok":"tiktok","Instagram 피드":"ig_feed","Instagram 릴스":"ig_reels",
                       "YouTube 숏츠":"youtube","YouTube":"youtube"}
                    return m.get(stated,"tiktok")
                url=str(url).lower()
                if "tiktok.com" in url: return "tiktok"
                if "instagram.com" in url: return "ig_reels" if "/reel" in url else "ig_feed"
                if "youtube.com" in url: return "youtube"
                return "tiktok"

            results_by_platform = {p:[] for p in PLATFORMS}
            scraped_raw = {p:[] for p in PLATFORMS}
            total_n=len(df_kol)
            prog=st.progress(0,"준비 중...")

            for i,row in df_kol.iterrows():
                name=str(row.get("KOL명","") or f"KOL {i+1}")
                url=str(row.get("URL (필수)",""))
                cost=float(row.get("캐스팅 비용(JPY)",0) or 0)
                stated_plat=str(row.get("플랫폼","") or "")
                plat=detect_plat(url,stated_plat)

                prog.progress((i+0.3)/total_n, f"처리 중: {name} ({plat})")

                if test_mode:
                    if plat=="tiktok":
                        data={"views":_r.randint(50000,900000),"likes":_r.randint(1000,40000),
                              "comments":_r.randint(20,500),"saves":_r.randint(100,10000),
                              "shares":_r.randint(10,500),"posts_scraped":n_posts,"error":None}
                    elif plat in ("ig_feed","ig_reels"):
                        data={"likes":_r.randint(500,15000),"comments":_r.randint(5,200),
                              "views":_r.randint(10000,200000),"posts_scraped":n_posts,"error":None}
                    else:
                        data={"views":_r.randint(100000,5000000),"likes":_r.randint(1000,50000),
                              "comments":_r.randint(10,500),"posts_scraped":n_posts,"error":None}
                else:
                    try:
                        from scraper import scrape_tiktok, scrape_instagram, scrape_youtube
                        tk_sid=st.session_state.sessions.get("tiktok","") or None
                        ig_sid=st.session_state.sessions.get("instagram","") or None
                        if plat=="tiktok":
                            data=scrape_tiktok(url,n_posts,session_id=tk_sid)
                        elif plat in ("ig_feed","ig_reels"):
                            ct="feed" if plat=="ig_feed" else "reels"
                            data=scrape_instagram(url,n_posts,ct,session_id=ig_sid)
                        else:
                            data=scrape_youtube(url,n_posts)
                    except Exception as e:
                        data={"error":str(e),"posts_scraped":0}

                if data.get("error"):
                    st.warning(f"⚠️ {name}: {data['error']}")

                scraped_raw[plat].append({"name":name,"cost_jpy":cost,**data})
                prog.progress((i+1)/total_n, f"완료: {name}")

            prog.progress(1.0,"✅ 스크래핑 완료!")

            # 평가
            for plat,raw_list in scraped_raw.items():
                valid=[d for d in raw_list if not d.get("error")]
                if not valid: continue
                if plat=="tiktok":
                    filt=[d for d in valid if d.get("views",0)>0]
                else:
                    filt=[d for d in valid if (d.get("likes",0)+d.get("comments",0))>0]
                if filt:
                    results_by_platform[plat]=evaluate_batch(filt,plat if plat!="ig_reels" else "ig_feed")

            st.session_state["scrape_results"]=results_by_platform

    if "scrape_results" in st.session_state:
        show_results(st.session_state["scrape_results"])


# ─── 수동 입력 ────────────────────────────────────────────
with tab_manual:
    st.markdown('<div class="sh">📝 수동 데이터 입력 & 즉시 평가</div>', unsafe_allow_html=True)

    plat_tabs=st.tabs(["🎵 TikTok","📸 IG 피드","🎬 IG 릴스"])

    manual_results={}

    for plat,tab,icon in [("tiktok",plat_tabs[0],"🎵"),("ig_feed",plat_tabs[1],"📸"),("ig_reels",plat_tabs[2],"🎬")]:
        with tab:
            key_prefix=plat
            if f"{key_prefix}_list" not in st.session_state:
                st.session_state[f"{key_prefix}_list"]=[{"name":"KOL 1","cost":500000}]

            b1,b2=st.columns(2)
            if b1.button("➕ KOL 추가",key=f"{key_prefix}_add"):
                n=len(st.session_state[f"{key_prefix}_list"])+1
                st.session_state[f"{key_prefix}_list"].append({"name":f"KOL {n}","cost":500000})
            if b2.button("🗑️ 초기화",key=f"{key_prefix}_clr"):
                st.session_state[f"{key_prefix}_list"]=[{"name":"KOL 1","cost":500000}]

            inputs=[]
            for idx,kol in enumerate(st.session_state[f"{key_prefix}_list"]):
                with st.expander(f"**{kol.get('name','KOL')}**",expanded=(idx==0)):
                    nc,cc=st.columns(2)
                    name=nc.text_input("KOL명",value=kol.get("name",""),key=f"{key_prefix}_n_{idx}")
                    cost=cc.number_input("비용",0,value=int(kol.get("cost",500000)),step=10000,key=f"{key_prefix}_c_{idx}")
                    st.session_state[f"{key_prefix}_list"][idx].update({"name":name,"cost":cost})

                    if plat=="tiktok":
                        a,b,c_=st.columns(3)
                        views=a.number_input("조회수",0,value=int(kol.get("views",0)),key=f"{key_prefix}_v_{idx}")
                        likes=b.number_input("좋아요",0,value=int(kol.get("likes",0)),key=f"{key_prefix}_l_{idx}")
                        cmts=c_.number_input("댓글",0,value=int(kol.get("cmts",0)),key=f"{key_prefix}_co_{idx}")
                        d,e=st.columns(2)
                        saves=d.number_input("저장",0,value=int(kol.get("saves",0)),key=f"{key_prefix}_s_{idx}")
                        shares=e.number_input("공유 ★",0,value=int(kol.get("shares",0)),key=f"{key_prefix}_sh_{idx}")
                        st.session_state[f"{key_prefix}_list"][idx].update({"views":views,"likes":likes,"cmts":cmts,"saves":saves,"shares":shares})
                        inputs.append({"name":name,"cost_jpy":cost,"views":views,"likes":likes,"comments":cmts,"saves":saves,"shares":shares})
                    elif plat=="ig_feed":
                        a,b=st.columns(2)
                        likes=a.number_input("좋아요",0,value=int(kol.get("likes",0)),key=f"{key_prefix}_l_{idx}")
                        cmts=b.number_input("댓글",0,value=int(kol.get("cmts",0)),key=f"{key_prefix}_co_{idx}")
                        st.session_state[f"{key_prefix}_list"][idx].update({"likes":likes,"cmts":cmts})
                        inputs.append({"name":name,"cost_jpy":cost,"likes":likes,"comments":cmts})
                    else:  # ig_reels
                        a,b,c_=st.columns(3)
                        views=a.number_input("조회수",0,value=int(kol.get("views",0)),key=f"{key_prefix}_v_{idx}")
                        likes=b.number_input("좋아요",0,value=int(kol.get("likes",0)),key=f"{key_prefix}_l_{idx}")
                        cmts=c_.number_input("댓글",0,value=int(kol.get("cmts",0)),key=f"{key_prefix}_co_{idx}")
                        st.session_state[f"{key_prefix}_list"][idx].update({"views":views,"likes":likes,"cmts":cmts})
                        inputs.append({"name":name,"cost_jpy":cost,"views":views,"likes":likes,"comments":cmts})

                    if len(st.session_state[f"{key_prefix}_list"])>1 and st.button("❌ 삭제",key=f"{key_prefix}_del_{idx}"):
                        st.session_state[f"{key_prefix}_list"].pop(idx); st.rerun()

            if st.button(f"▶ {PLATFORMS.get(plat,{}).get('label',plat)} 평가 실행",type="primary",use_container_width=True,key=f"{key_prefix}_run"):
                if plat=="tiktok":
                    valid=[k for k in inputs if k["cost_jpy"]>0 and k["views"]>0]
                else:
                    valid=[k for k in inputs if k["cost_jpy"]>0 and (k["likes"]+k["comments"])>0]
                if valid:
                    eval_plat="ig_feed" if plat in ("ig_feed","ig_reels") else plat
                    manual_results[plat]=evaluate_batch(valid,eval_plat)
                    st.session_state["manual_results"]=manual_results
                else:
                    st.warning("유효한 데이터를 입력하세요.")

    if "manual_results" in st.session_state and st.session_state["manual_results"]:
        show_results(st.session_state["manual_results"])


# ─── 가이드 ──────────────────────────────────────────────
with tab_guide:
    st.markdown('<div class="sh">📖 KOL 선별 프레임워크 가이드 v4</div>', unsafe_allow_html=True)

    with st.expander("플랫폼별 가중치",expanded=True):
        c1,c2,c3,c4=st.columns(4)
        c1.markdown("**TikTok**")
        c1.table(pd.DataFrame({"지표":["CPV","ER%","저장률","공유율★"],"가중치":["30%","30%","25%","15%"]}))
        c2.markdown("**IG 릴스**")
        c2.table(pd.DataFrame({"지표":["CPV","ER%","댓글%"],"가중치":["30%","35%","35%"]}))
        c3.markdown("**IG 피드**")
        c3.table(pd.DataFrame({"지표":["CPE","댓글·좋아요비율"],"가중치":["60%","40%"]}))
        c4.markdown("**YouTube**")
        c4.table(pd.DataFrame({"지표":["CPV"],"가중치":["100%"]}))

    with st.expander("TikTok 컷오프"):
        st.table(pd.DataFrame({
            "지표":["CPV(₩)","ER%","저장률%","공유율%★"],
            "탁월(E)":["<12","≥3.48%","≥0.490%","≥0.058%"],
            "양호(G)":["12~23","1.85~3.48%","0.230~0.490%","0.024~0.058%"],
            "보통(C)":["23~29","1.05~1.85%","0.122~0.230%","0.013~0.024%"],
            "저조(H)":["≥29","<1.05%","<0.122%","<0.013%"],
        }).set_index("지표"))

    with st.expander("Instagram 피드 컷오프"):
        st.table(pd.DataFrame({
            "지표":["CPE(₩)","댓글·좋아요비율"],
            "탁월(E)":["<1,631","≥0.0042"],
            "양호(G)":["1,631~3,465","0.0000~0.0042"],
            "보통(C)":["3,465~6,261","~0.0000"],
            "저조(H)":["≥6,261","<0.0000"],
        }).set_index("지표"))

    with st.expander("결론 로직 v4"):
        st.markdown("""
1. **1차 필터:** 절대점수 < 5.0 → 제외 ✗
2. **예외:** 절대 ≥ 8.0 & 상대 < 4.0 → 검토 △ + 코멘트
3. **2차 (고정 임계값):** 상대 ≥ 6.0 → 채택 ✓ | 4.0~5.9 → 검토 △ | < 4.0 → 제외 ✗
        """)
        cols=st.columns(3)
        cols[0].markdown('<div style="background:#1D6A2D;color:#fff;padding:10px;border-radius:8px;text-align:center;font-weight:700">채택 ✓<br><small>상대 ≥ 6.0</small></div>',unsafe_allow_html=True)
        cols[1].markdown('<div style="background:#DCB306;color:#fff;padding:10px;border-radius:8px;text-align:center;font-weight:700">검토 △<br><small>4.0~5.9</small></div>',unsafe_allow_html=True)
        cols[2].markdown('<div style="background:#7B0000;color:#fff;padding:10px;border-radius:8px;text-align:center;font-weight:700">제외 ✗<br><small>< 4.0</small></div>',unsafe_allow_html=True)
